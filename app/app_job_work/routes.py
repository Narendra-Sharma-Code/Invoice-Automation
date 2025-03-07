from flask import Flask, send_file,jsonify
from flask import Blueprint, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side   
from openpyxl import load_workbook
import re
import os
import pandas as pd
from . import constants
from num2words import num2words
from app.extensions import mysql  # Import mysql from extensions
import json
from openpyxl.utils import get_column_letter,column_index_from_string
import math
import difflib
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

# app_job_work = Flask(__name__)
app_job_work = Blueprint('app_job_work', __name__, template_folder='templates',static_folder='app/static')
    
# Ensure the output directory exists
OUTPUT_DIR = os.path.join(os.path.dirname(__file__),"outputs")
os.makedirs(OUTPUT_DIR, exist_ok=True)


# Styles for the Excel file
BOLD_FONT = Font(bold=True, size=12)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal="right",vertical='top', wrap_text=True)
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
DARK_BOTTOM_BORDER = Border(bottom=Side(style='medium'))
DARK_VERTICAL_BORDER = Border(right=Side(style='medium'))
DARK_BOTTOM_EXTENDED_BORDER = Border(bottom=Side(style='medium'), right=Side(style='medium'))
VERTICAL_BOLD_BORDER = Border(left=Side(style='medium'))  # For bold vertical borders
WRAP_TEXT_ALIGN = Alignment(wrap_text=True)  # To wrap the paragraph text

# Helper function to set cell value, style, and borders
def set_cell(ws, cell, value, font=None, alignment=None, border=None):
    ws[cell] = value
    if font:
        ws[cell].font = font
    if alignment:
        ws[cell].alignment = alignment
    if border:
        ws[cell].border = border

@app_job_work.route('/upload_job_work')
def upload_job_work():
    return render_template('upload_job_work.html') # HTML template for file upload

    
@app_job_work.route('/process', methods=['POST'])
def process_file():
        # Check if an input file was uploaded
    if 'file' not in request.files:
        return "No file uploaded", 400

    input_file = request.files['file']
    if input_file.filename == '':
        return "No selected file", 400

    # Load the input Excel file
    wb_input = load_workbook(input_file)
    ws_input = wb_input.active  # Assuming data is in the first sheet

    # Find the "Metal KT" column and collect unique data
    metal_kt_data = []
    metal_kt_index = None
    for row in ws_input.iter_rows(min_row=1, max_row=ws_input.max_row, values_only=True):
        if "Metal KT" in row and metal_kt_index is None:
            metal_kt_index = row.index("Metal KT")  # Find the index of "Metal KT" column
            continue  # Skip the header row
        if metal_kt_index is not None:
            value = row[metal_kt_index]
            if value and value not in metal_kt_data:  # Avoid empty cells and duplicates
                metal_kt_data.append(value)

    # Generate formatted headers, ensuring "Net" and "Pure" columns are side-by-side

    paired_headers = []
    other_headers = []
    fixed_headers = []
    silver_present = False
    platinum_present = False

    for item in metal_kt_data:
        if isinstance(item, str):
            if re.search(r'\d+(kt|ekt)', item, re.IGNORECASE):  # For Gold entries like "18KT" or "22EKT"
                number = re.search(r'\d+', item).group()  # Extract the number before "kt" or "ekt"
                if "ekt" in item.lower():
                    net_header = f"Net Wt (gms) {number}EKT Gold"
                    pure_header = f"Pure Wt (gms) {number}EKT Gold"
                else:
                    net_header = f"Net Wt (gms) {number}KT Gold"
                    pure_header = f"Pure Wt (gms) {number}KT Gold"
                paired_headers.extend([net_header, pure_header])
            elif "S.S" in item or "Silver" in item:  # For "S.S" or "Silver"
                silver_present = True
            elif re.search(r'\b(platinum|plt)\b', item, re.IGNORECASE):  # For "Platinum" or "plt"
                platinum_present = True
                
            else:
            # Capture any other unique Metal KT values for inclusion as headers
                other_headers.append(item)

# Add the silver column if needed
    if silver_present:
        paired_headers.extend(["Net Wt (gms) Silver", "Pure Wt (gms) Silver"])  # Ensure side-by-side placement

# Add the platinum column if needed
    if platinum_present:
        paired_headers.extend(["Net Wt (gms) Platinum", "Pure Wt (gms) Platinum"])  # Ensure side-by-side placement

# Combine all headers
    all_headers = fixed_headers + paired_headers + other_headers

    # Create a new Excel workbook and active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # Setting up the heading
    ws.merge_cells('A1:D1')
    set_cell(ws, 'A1', "INVOICE", font=BOLD_FONT, alignment=CENTER_ALIGN)

    # Create a horizontal line between row 31 and 32 from column A to P
    for col in range(1, 17):  # Columns A (1) to P (16)
        cell = ws.cell(row=31, column=col)
        cell.border = DARK_BOTTOM_BORDER

    # Add "Banker :" in row 26, column 5 in bold letters
    set_cell(ws, 'E26', "Banker :", font=BOLD_FONT, alignment=LEFT_ALIGN)
    
    # Create a horizontal line between row 26 and 27 from column D to P
    for col in range(4, 17):  # Columns D (4) to P (16)
        cell = ws.cell(row=26, column=col)
        cell.border = DARK_BOTTOM_BORDER

    # Create a horizontal line between row 30 and 31 from column A to P
    for col in range(1, 17):  # Columns A (1) to P (16)
        cell = ws.cell(row=30, column=col)
        cell.border = DARK_BOTTOM_BORDER


    # Adding a horizontal line between row 20 and 21 from column D to P
    for col in range(4, 17):  # Columns D (4) to P (16)
        cell = ws.cell(row=20, column=col)
        cell.border = DARK_BOTTOM_BORDER

    # Adding "Country of Origin Of Goods :" in row 21, columns E, F, G in bold letters
    ws.merge_cells('E21:G21')  # Merge cells E21, F21, and G21
    set_cell(ws, 'E21', "Country of Origin Of Goods :", font=BOLD_FONT, alignment=LEFT_ALIGN)

    # Adding "INDIA" in row 22, column E without bold letters
    set_cell(ws, 'E22', "INDIA", font=None, alignment=LEFT_ALIGN)
    
    # Adding "Bank Details :" in column 5 (E), row 16, in bold letters
    set_cell(ws, 'E16', "Bank Details :", font=BOLD_FONT, alignment=LEFT_ALIGN)

    # Adding a bold bottom border between row 14 and 15, from column D to P
    for col in range(4, 17):  # Columns D (4) to P (16)
        cell = ws.cell(row=14, column=col)
        cell.border = DARK_BOTTOM_BORDER

   # Merge columns 5, 6, and 7 in row 10
    ws.merge_cells('E10:F10')
    set_cell(ws, 'E10', "Buyer's Ord No. & Date Ref.:", font=BOLD_FONT, alignment=LEFT_ALIGN)

    buyers_ord_no_date = request.form.get('buyers_ord_no_date')

# Merge cells for the actual invoice value (entered from the form)
    ws.merge_cells('G10:H10')
    set_cell(ws, 'G10', buyers_ord_no_date, alignment=LEFT_ALIGN)

    # Merge columns E, F, and G in row 11
    ws.merge_cells('E11:G11')  # Columns E, F, and G merged
    set_cell(ws, 'E11', "Buyer (if other than consignee) :", alignment=LEFT_ALIGN)

    # Merge cells for "Invoice No. & Date :" label
    ws.merge_cells('E2:F2')
    set_cell(ws, 'E2', "Invoice No. & Date :", font=BOLD_FONT, alignment=LEFT_ALIGN)

    invoice_no_date = request.form.get('invoice_no_date')

# Merge cells for the actual invoice value (entered from the form)
    ws.merge_cells('E3:F3')
    set_cell(ws, 'E3', invoice_no_date, alignment=LEFT_ALIGN)

# Merge cells for "Exporter's Ref:" label
    ws.merge_cells('I2:J2')
    set_cell(ws, 'I2', "Exporter's Ref:", font=BOLD_FONT, alignment=LEFT_ALIGN)

    Exporter_Ref = request.form.get('Exporter_Ref')

# Merge cells for the actual Exporter's Ref value (entered from the form)
    ws.merge_cells('I3:J3')
    set_cell(ws, 'I3', Exporter_Ref,alignment=LEFT_ALIGN)

    # Adding a horizontal bold bottom border between Row 8 and Row 9 from Column D to P
    for col in range(4, 17):  # Columns D (4) to P (16)
        cell = ws.cell(row=8, column=col)
        cell.border = Border(bottom=Side(style='medium'))  # Apply bold bottom border


    # Adding a vertical bold line between Columns D and E for Rows 2 to 31
    for row in range(2, 32):  # Rows 2 to 31
        ws.cell(row=row, column=4).border = Border(right=Side(style='medium'))  # Bold right border for Column D


    # Exporter Section (Heading Only)
    ws.merge_cells('A2:D2')
    set_cell(ws, 'A2', "Exporter :", font=BOLD_FONT, alignment=LEFT_ALIGN)

    # Factory Address Section
    ws.merge_cells('A3:B7')
    set_cell(ws, 'A3', "Factory Address :", font=BOLD_FONT, alignment=LEFT_ALIGN)

    ws.merge_cells('C3:D7')
    set_cell(ws, 'C3', ("M/S. UNI DESIGN ELITE JEWELLERY PVT LTD\n"
                        "Survey No.280091, Mahendra Brothers Exports Pvt Ltd\n"
                        "Gandevi Road, Jamalpore\n"
                        "At Village Navasari, At Taluka Navasari, At District Navasari\n"
                        "Pin No.396445\n"
                        "IGST NO:24AAACK3499E1ZL"), alignment=LEFT_ALIGN)

    # Sales Office Address Section
    ws.merge_cells('A9:B9')
    set_cell(ws, 'A9', "Sales Office Address :", font=BOLD_FONT, alignment=LEFT_ALIGN)

    ws.merge_cells('C9:D14')
    set_cell(ws, 'C9', ("Unidesign Elite Jewellery Pvt Ltd\n"
                        "UNI DESIGN ELITE JEWELLERY PVT LTD\n"
                        "PLOT NO D-7/1, 1ST FLOOR, ASIAN HOUSE, ROAD NO 16\n"
                        "OPP. PRASAD LAB, MIDC ANDHERI -(E), M-93\n"
                        "Tel : 30815888\n"
                        "GST NO: 27AAACK3499E2ZE\n"
                        "LUT ARN No- AD270324023939U\n"
                        "Date: 01/4/2024 To 31/03/2025"), alignment=LEFT_ALIGN)

    # Adding a dark bottom border for row 14
    for col in range(1, 5):  # Columns A to D
        cell = ws.cell(row=14, column=col)
        cell.border = DARK_BOTTOM_BORDER

    # Consignee Section
    ws.merge_cells('A16:B16')
    set_cell(ws, 'A16', "Consignee :", font=BOLD_FONT, alignment=LEFT_ALIGN)

    ws.merge_cells('C16:D22')
    set_cell(ws, 'C16', ("UNIDESIGN JEWELLERY PVT.LTD. UNIT III\n"
                         "PLOT # 4, 5 & 6 (Part), Seepz, Sez,\n"
                         "Andheri (East), Mumbai-400096\n\n"
                         "Tel : 022 6668 1050\n"
                         "Fax : GST - 27AAACU0572GIZH\n"
                         "GST : 27AAACU0572G1ZH\n"
                         "Pan No. : AAACU0572G"), alignment=LEFT_ALIGN)

    # Adding a dark bottom border for row 22
    for col in range(1, 5):  # Columns A to D
        cell = ws.cell(row=22, column=col)
        cell.border = DARK_BOTTOM_BORDER

    # Pre-Carriage Section
    ws.merge_cells('A24:B24')
    set_cell(ws, 'A24', "Pre-Carriage by :", font=BOLD_FONT, alignment=LEFT_ALIGN)

    # Adding "Place of Receipt by Pre-Carrier :" in column 3, row 24
    set_cell(ws, 'C24', "Place of Receipt by Pre-Carrier :", font=BOLD_FONT, alignment=LEFT_ALIGN)

    # Adding "N.A" in column A just below "Pre-Carriage by :"
    set_cell(ws, 'A25', "N.A", alignment=LEFT_ALIGN)

    # Adding "N.A." in column 3, row 25
    set_cell(ws, 'C25', "N.A.", alignment=LEFT_ALIGN)

    # Adding "Port of Loading :" in column 3, row 26
    set_cell(ws, 'C26', "Port of Loading :", font=BOLD_FONT, alignment=LEFT_ALIGN)

    # Adding "Final Destination :" in column 3, row 27
    set_cell(ws, 'C27', "Final Destination :", font=BOLD_FONT, alignment=LEFT_ALIGN)

    # Adding "INDIA " in column 3, row 28
    set_cell(ws, 'C28', "INDIA", alignment=LEFT_ALIGN)

    # Adding extended bold bottom border for rows 25 and 26
    for col in range(1, 3):  # Columns A to C
        cell = ws.cell(row=25, column=col)
        cell.border = DARK_BOTTOM_EXTENDED_BORDER

    # Adding "Vessel/Flight No. :" in row 26
    set_cell(ws, 'A26', "Vessel/Flight No. :", font=BOLD_FONT, alignment=LEFT_ALIGN)

    # Adding extended bold bottom border for rows 26 and 27
    for col in range(1, 3):  # Columns A to C
        cell = ws.cell(row=26, column=col)
        cell.border = DARK_BOTTOM_EXTENDED_BORDER

    # Adding "Port of Discharge :" in row 27
    set_cell(ws, 'A27', "Port of Discharge :", font=BOLD_FONT, alignment=LEFT_ALIGN)

    # Adding bold vertical border between columns B and D for rows 22 to 30
    for row in range(23, 31):  # Rows 22 to 30
        ws.cell(row=row, column=3).border = VERTICAL_BOLD_BORDER  # Bold vertical line between columns B

    # Add a horizontal line between row 22 and 23 from columns D to P
    for col in range(4, 17):  # Columns D to P (4 to 16)
        cell = ws.cell(row=22, column=col)
        cell.border = Border(bottom=Side(style='medium'))
    
    # Add a horizontal line between row 1 and 2 from columns D to P
    for col in range(4, 17):  # Columns D to P (4 to 16)
        cell = ws.cell(row=1, column=col)
        cell.border = Border(bottom=Side(style='medium'))

    # Adding horizontal line (bottom border) between row 25 and 26 from column B to E
    for col in range(2, 5):  # Columns B to E
        cell = ws.cell(row=25, column=col)
        cell.border = Border(bottom=Side(style='medium'))

    for col in range(2, 5):  # Columns B to E
        cell = ws.cell(row=26, column=col)
        cell.border = Border(bottom=Side(style='medium'))
    
    # Adding "Terms of Delivery & Payment :" in row 23, columns E, F, and G in bold letters
    ws.merge_cells('E23:G23')  # Merge columns E, F, and G in row 23
    set_cell(ws, 'E23', "Terms of Delivery & :", font=BOLD_FONT, alignment=LEFT_ALIGN)

# Adding "Payment Term: Immediate" in row 24, columns E, F, and G
    ws.merge_cells('E24:G24')  # Merge columns E, F, and G in row 24
    set_cell(ws, 'E24', "Payment Term: Immediate", font=None, alignment=LEFT_ALIGN)

# Create a horizontal line between row 25 and 26 from column D to P
    for col in range(4, 17):  # Columns D to P (4 to 16)
        cell = ws.cell(row=25, column=col)
        cell.border = Border(bottom=Side(style='medium'))
        
    # Adding Metal KT data as a table in the output Excel
    start_row = ws.max_row + 2  # Start below the existing formatted data

    # First column: "Marks & Nos./ Container No."
    set_cell(ws, f'A{start_row}', "Marks & Nos./ Container No.", font=BOLD_FONT, alignment=CENTER_ALIGN)
    set_cell(ws, f'B{start_row}', "No. & Kind of Pkgs", font=BOLD_FONT, alignment=CENTER_ALIGN)
    # set_cell(ws, f'C{start_row}', "Description of Goods", font=BOLD_FONT, alignment=CENTER_ALIGN)
    
    from openpyxl.styles import Font, Alignment

    # Set the "Description of Goods" header at C{start_row}
    set_cell(ws, f'C{start_row}', "Description of Goods", font=BOLD_FONT, alignment=CENTER_ALIGN)

# Print "71131913" just above the "Description of Goods"
    description_goods_row = start_row - 1  # Row immediately above "Description of Goods"
    ws[f'C{description_goods_row}'] = "71131913"  # Print the value
    ws[f'C{description_goods_row}'].font = Font(size=10, color="808080")  # Light font (gray) and smaller font size
    ws[f'C{description_goods_row}'].alignment = Alignment(horizontal='right', vertical='top', wrap_text=False)  # Right-aligned text

    
    set_cell(ws, f'D{start_row}', "Gross Wt (gms)", font=BOLD_FONT, alignment=CENTER_ALIGN)
    set_cell(ws, f'E{start_row}', "pure wt (gms) 0.995 gold", font=BOLD_FONT, alignment=CENTER_ALIGN)
    
    
   # Adding all Metal KT data headers
    current_col = 6  # Start from column E
    for header in all_headers:  # all_headers contains Metal KT-related columns
        col_letter = ws.cell(row=start_row, column=current_col).column_letter
        set_cell(ws, f'{col_letter}{start_row}', header, font=BOLD_FONT, alignment=CENTER_ALIGN)
        current_col += 1

    # Adding "Qty. Pcs" column
    qty_pcs_col_letter = ws.cell(row=start_row, column=current_col).column_letter
    set_cell(ws, f'{qty_pcs_col_letter}{start_row}', "Qty. Pcs", font=BOLD_FONT, alignment=CENTER_ALIGN)
    current_col += 1  # Increment column index for the next column

    # Adding "Rate Av. Per Pc" column
    rate_av_col_letter = ws.cell(row=start_row, column=current_col).column_letter
    set_cell(ws, f'{rate_av_col_letter}{start_row}', "Rate Av. Per Pc", font=BOLD_FONT, alignment=CENTER_ALIGN)
    current_col += 1  # Increment column index for any subsequent additions

    # Adding "Labor Amt" column
    labor_amt_col_letter = ws.cell(row=start_row, column=current_col).column_letter
    set_cell(ws, f'{labor_amt_col_letter}{start_row}', "Labour Amt", font=BOLD_FONT, alignment=CENTER_ALIGN)
    current_col += 1  # Increment column index for any subsequent additions
    
    # Adding "Metal Amt." column
    metal_amt_col_letter = ws.cell(row=start_row, column=current_col).column_letter
    set_cell(ws, f'{metal_amt_col_letter}{start_row}', "Metal Amt.", font=BOLD_FONT, alignment=CENTER_ALIGN)
    current_col += 1  # Increment column index for any subsequent additions
    
    # Adding "Diamond CTS" column after "Metal Amt."
    diamond_cts_col_letter = ws.cell(row=start_row, column=current_col).column_letter
    set_cell(ws, f'{diamond_cts_col_letter}{start_row}', "Diamond CTS", font=BOLD_FONT, alignment=CENTER_ALIGN)
    current_col += 1  # Increment column index for the next column
    
   # Find the position of "Diamond CTS" column dynamically
    diamond_cts_col = None
    for col_idx in range(1, ws.max_column + 1):
        if ws.cell(row=start_row, column=col_idx).value == "Diamond CTS":
            diamond_cts_col = col_idx
            break

# Add "Diamond Amt." column immediately after "Diamond CTS"
    if diamond_cts_col is not None:
    # Calculate the column letter after "Diamond CTS"
        diamond_amt_col_letter = ws.cell(row=start_row, column=diamond_cts_col + 1).column_letter
        set_cell(ws, f'{diamond_amt_col_letter}{start_row}', "Diamond Amt.", font=BOLD_FONT, alignment=CENTER_ALIGN)

    # Find the position of "Diamond Amt." column dynamically
    diamond_amt_col = None
    for col_idx in range(1, ws.max_column + 1):
        if ws.cell(row=start_row, column=col_idx).value == "Diamond Amt.":
            diamond_amt_col = col_idx
            break

# Add "Color Stone CTS" column immediately after "Diamond Amt."
    if diamond_amt_col is not None:
    # Calculate the column letter after "Diamond Amt."
        color_stone_cts_col_letter = ws.cell(row=start_row, column=diamond_amt_col + 1).column_letter
        set_cell(ws, f'{color_stone_cts_col_letter}{start_row}', "Color Stone CTS", font=BOLD_FONT, alignment=CENTER_ALIGN)

    # Find the position of "Color Stone CTS" column dynamically
    color_stone_cts_col = None
    for col_idx in range(1, ws.max_column + 1):
        if ws.cell(row=start_row, column=col_idx).value == "Color Stone CTS":
            color_stone_cts_col = col_idx
            break

# Add "Color Stone Amt." column immediately after "Color Stone CTS."
    if color_stone_cts_col is not None:
    # Calculate the column letter after "Color Stone CTS."
        color_stone_amt_col_letter = ws.cell(row=start_row, column=color_stone_cts_col + 1).column_letter
        set_cell(ws, f'{color_stone_amt_col_letter}{start_row}', "Color Stone Amt.", font=BOLD_FONT, alignment=CENTER_ALIGN)
    # Add "Amount US$" as the last column in the output Excel file
    last_col_index = ws.max_column + 1  # Determine the next available column
    last_col_letter = ws.cell(row=start_row, column=last_col_index).column_letter
    set_cell(ws, f'{last_col_letter}{start_row}', "Amount US$", font=BOLD_FONT, alignment=CENTER_ALIGN)
   
    # # Add "0.995 Gold Pure Wt" as the next column after "Amount US$"
    # gold_pure_wt_col_index = ws.max_column + 1  # Increment by 1 for the next column
    # gold_pure_wt_col_letter = ws.cell(row=start_row, column=gold_pure_wt_col_index).column_letter
    # set_cell(ws, f'{gold_pure_wt_col_letter}{start_row}', "pure wt (gms) 0.995 gold", font=BOLD_FONT, alignment=CENTER_ALIGN)

    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    rm_list = request.form.getlist('rm[]')
    

    # Adjust widths for dynamically added columns
    for col_index in range(5, ws.max_column + 1):  # From the first added column to the last
        col_letter = ws.cell(row=start_row, column=col_index).column_letter
        ws.column_dimensions[col_letter].width = 15  # Adjust as needed for clarity
    
    
    output_header = [cell.value.strip() for cell in ws[33] if cell.value]
   
  # Load the parsed data
    df = pd.read_excel(input_file, header=3)

# Step 1: Clean column names
    df.columns = df.columns.str.strip()  # Remove leading/trailing spaces
    df.columns = df.columns.str.replace(r'\s+', ' ', regex=True)  # Replace multiple spaces with a single space
    df.columns = df.columns.str.replace(r'[^A-Za-z0-9 ]+', '', regex=True)  # Remove special characters
 
# Step 2: Add default values for missing columns
    required_columns = [
        'INR LABOUR', 'Inv Rm Wt', 'Design', 'Ctg', 'Metal KT', 
        'Gross Wt', 'DsgCtg', 'Karatage', 'Inv Pure Wt', 
        'Inv Rate', 'Inv Value', 'Labour', 'Qty','Rm Code'
    ]
    
    


# Check for missing columns and add them with default value of 0
    for col in required_columns:
        if col not in df.columns:
            df[col] = 0  # Add the missing column with default value 0

# Step 3: Perform aggregation
    grouped = df.groupby(["Design", "Ctg", "Metal KT"], dropna=False).agg({
        "Gross Wt": "sum",
        "DsgCtg": "first",
        "Karatage": "first",
        "Inv Rm Wt": "sum",
        "Inv Pure Wt": "sum",
        "Inv Rate": "mean",  # Average rate
        "Inv Value": "sum",
        "Labour": "sum",
        "INR LABOUR": "sum",
        "Qty": "sum",
        "Rm Code": "first"  # Include Rm Code
    }).reset_index()

    
    final_list = []
    prev = None
    gross_wt = qty = labour = rate_avg_pp = metal_amt = total_usd = stone_amt = 0
    design_ctg = product_name = ''
    ctg = []
    rm_code_list = []  # New list for Rm Code values
    temp_dict = {}
    metal_temp = {}
    

    for index, row in grouped.iterrows():
        if not prev:
            prev = row['Design']
        elif prev != row['Design'] or index == (grouped.shape[0] - 1):
            if index == (grouped.shape[0] - 1):
                ctg.append(f"{row['Ctg']},{row['Metal KT']}")
                # Also collect the Rm Code from the row
                rm_code_list.append(str(row['Rm Code']).strip())
                design_ctg = row['DsgCtg']
                gross_wt += row['Gross Wt']
                qty += row['Qty']
                labour += row['Labour']
                metal_amt += row['Inv Value'] if row['Ctg'] in ['G','M','S'] else 0
                stone_amt += row['Inv Value'] if row['Ctg'] in ['C','D'] else 0
                metal_temp.update(get_metal_mapping(output_header, row, row['Ctg'], row['Metal KT']))

            ctg_str = ", ".join(ctg)
            rm_code_str = ", ".join(rm_code_list)
            total_usd = labour + metal_amt + stone_amt
            rate_avg_pp = total_usd / qty if qty != 0 else rate_avg_pp
            product_name = get_product_name(ctg_str, rm_code_str)
            design_ctg = (constants.DESIGN_CATEGORY[design_ctg] if design_ctg in constants.DESIGN_CATEGORY else '')

            temp_dict = {
                'Marks & Nos./ Container No.': f'{product_name},\r\n {round(qty)}, {design_ctg}',
                'No. & Kind of Pkgs': '',
                'Description of Goods': '',
                'Design': prev,
                'Gross Wt (gms)': f"{round(gross_wt, 3):.3f}",
                'Qty. Pcs': round(qty),
                'Labour Amt': round(labour, 2),
                'Metal Amt.': round(metal_amt, 2),
                'Rate Av. Per Pc': round(rate_avg_pp, 2),
                'Amount US$': round(total_usd, 2)
            }

            temp_dict.update(metal_temp)
            final_list.append(temp_dict)

        # Reset variables for the next Design
            gross_wt = qty = labour = rate_avg_pp = metal_amt = total_usd = stone_amt = 0
            design_ctg = product_name = ''
            ctg = []
            rm_code_list = []  # Reset the Rm Code list
            temp_dict = {}
            metal_temp = {}
            prev = row['Design']

        ctg.append(f"{row['Ctg']},{row['Metal KT']}")
        rm_code_list.append(str(row['Rm Code']).strip())
        design_ctg = row['DsgCtg']
        gross_wt += row['Gross Wt']
        qty += row['Qty']
        labour += row['Labour']
        metal_amt += row['Inv Value'] if row['Ctg'] in ['G','M','S'] else 0
        stone_amt += row['Inv Value'] if row['Ctg'] in ['C','D'] else 0
        metal_temp.update(get_metal_mapping(output_header, row, row['Ctg'], row['Metal KT']))

    # Map headers to the final output
    mapped_data = map_headers_to_data(output_header, final_list)
    # print(mapped_data)
    

    # Create final DataFrame from the mapped data
    final_output = pd.DataFrame(mapped_data)
    # print(final_output.to_string())
   
    # --- Existing code for creating your DataFrame ---
    mapped_data = map_headers_to_data(output_header, final_list)
    final_output = pd.DataFrame(mapped_data)

    # --- Debug: Print the rm_list and current DataFrame columns ---
    print("rm_list:", rm_list)
    print("DataFrame columns:", final_output.columns.tolist())

    # --- Normalize the rm_list values: strip spaces and lower-case ---
    normalized_rm_list = [item.strip().lower() for item in rm_list]

    # --- Check if "999pt" (normalized) is in the normalized_rm_list ---
    if "0.999 pt" in normalized_rm_list:
        print("'0.999 pt' found in rm_list (after normalization).")
        
        # --- Normalize the DataFrame column names for the check ---
        # Create a dictionary mapping normalized column names to the actual column names
        col_map = {col.strip().lower(): col for col in final_output.columns}
        
        # Check if a normalized column "950pt" exists
        if "950pt" in col_map:
            actual_col_950 = col_map["950pt"]
            print(f"Found '950pt' column as '{actual_col_950}' in final_output.")
            try:
                # Ensure the values in the "950pt" column are numeric.
                final_output[actual_col_950] = pd.to_numeric(final_output[actual_col_950], errors="coerce")
                
                # Calculate the new column values:
                # (value in "950pt") * 950 / 999
                pure_wt_999pt = final_output[actual_col_950] * 950 / 999

                # Find the index of the "950pt" column using the actual column name.
                index_950pt = final_output.columns.get_loc(actual_col_950)
                # Insert the new column "pt 0.999 pure wt" immediately after the "950pt" column.
                final_output.insert(index_950pt + 1, "pt 0.999 pure wt", pure_wt_999pt)

                print("Inserted 'pt 0.999 pure wt' column successfully.")
            except Exception as e:
                print("Error during calculation or column insertion:", e)
        else:
            print("Error: '950pt' column not found in final_output after normalization.")
    else:
        print("'999pt' is not present in rm_list (after normalization); skipping new column insertion.")

    # --- Debug: Print final columns to verify if the new column was added ---
    print("Final DataFrame columns:", final_output.columns.tolist())


    # Step 1: Clean the column names (strip spaces and convert to lowercase for consistency)
    final_output.columns = final_output.columns.str.strip().str.lower()  # Make column names lowercase and strip spaces

    # Step 2: Check if 'metal amt.' column exists (case insensitive)
    metal_amt_column_name = [col for col in final_output.columns if 'metal amt' in col.lower()]

    if metal_amt_column_name:
        # Step 3: Sum the 'Metal Amt.' column
        metal_amt_sum = final_output[metal_amt_column_name[0]].sum()  # Use the first matching column
        # print(f"Sum of '{metal_amt_column_name[0]}':", metal_amt_sum)
    else:
        print("Error: 'Metal Amt.' column not found.")
    


    # If mapped_data is a list of dictionaries or rows

    
    # Write the final DataFrame to the Excel file

    start_row = 34
    for j, header in enumerate(final_output.columns):
        ws.cell(row=start_row - 1, column=j + 1, value=header)

    for i, row in final_output.iterrows():
        for j, value in enumerate(row):
  
            ws.cell(row=start_row + i, column=j + 1, value=value) 
    last_row = start_row + len(final_output)  # Calculate the row where the mapping ends
    
  
    # Get multiple RM, QTY PCS, Met. Wt.Gms, and Value US$ from the form
    
    qty_pcs_list = request.form.getlist('qty_pcs[]')  
    met_wt_gms_list = request.form.getlist('met_wt_gms[]')
    value_usd_list = request.form.getlist('value_usd[]')
    rate_per_grams_list = request.form.getlist('rate[]')
   

    return_switch = request.form.get('return_switch', 'off')

    # Calculate starting rows
    header_row_for_return = last_row + 2
    content_row_for_return = header_row_for_return + 1

    
    # print(f"Return Switch Value: {return_switch}")  # Debug: Log received value

    if return_switch == "on":
        # print("Switch is ON. Executing logic...")

        # Add header text and make it bold
        ws[f'A{header_row_for_return}'] = "Balance Loose Metal Return"
        bold_font = Font(bold=True)
        ws[f'A{header_row_for_return}'].font = bold_font

        # Insert values from rm_list into column A
        for i, value_for_return in enumerate(rm_list, start=content_row_for_return):
            ws[f'A{i}'] = value_for_return

        # Dynamically adjust the next row based on inserted content
        content_end_row = content_row_for_return + len(rm_list)
    else:
        print("Switch is OFF or not set.")  # Debug: Log when switch is off
        # If switch is off, no data is added; use content_row_for_return directly
        content_end_row = content_row_for_return


    # Insert Headers for RM, QTY PCS, Met. Wt.Gms, and Value US$ with bold font
    headers_row_number = content_end_row + 3  # Dynamically adjust the header placement    
  
    # Create the headers
    ws.cell(row=headers_row_number, column=1, value="RM.")
    ws.cell(row=headers_row_number, column=1).font = Font(bold=True)  # Make the text bold

    # Check if any value is provided for QTY PCS
    qty_pcs_exists = any(qty for qty in qty_pcs_list if qty)  # Boolean flag to determine if QTY PCS exists
    
    if qty_pcs_exists:  # Print QTY PCS header only if at least one value exists
        ws.cell(row=headers_row_number, column=2, value="QTY PCS")
        ws.cell(row=headers_row_number, column=2).font = Font(bold=True)  
        met_wt_gms_col = 3  # Met. Wt.Gms will be in column 3
        value_usd_col = 4  # Value US$ will be in column 4
    else:
        met_wt_gms_col = 2  # Met. Wt.Gms will be in column 2 if QTY PCS does not exist
        value_usd_col = 3  # Value US$ will be in column 3 if QTY PCS does not exist
        
    ws.cell(row=headers_row_number, column=met_wt_gms_col, value="Met. Wt.Gms")
    ws.cell(row=headers_row_number, column=met_wt_gms_col).font = Font(bold=True)  # Make the text bold

    ws.cell(row=headers_row_number, column=value_usd_col, value="Value US$")
    ws.cell(row=headers_row_number, column=value_usd_col).font = Font(bold=True)  # Make the text bold
    
   
    # Step 4: Add Totals in the Last Row of Data
    data_start_row = headers_row_number + 1  # Row immediately after the headers
    last_data_row = data_start_row + len(rm_list)  # Get the last row of the data

    
    # Get the input value for Challan No.
    challan_no_value = request.form.get('challan_no', '')
    generated_table_data = []
    total_qty_pcs = 0
    total_met_wt_gms = 0
    total_value_usd = 0

    # Convert the generated table data into a JSON-compatible format
    for i in range(len(rm_list)):
        qty_pcs = float(qty_pcs_list[i]) if qty_pcs_list[i] else 0
        met_wt_gms = float(met_wt_gms_list[i]) if met_wt_gms_list[i] else 0
        value_usd = float(value_usd_list[i]) if value_usd_list[i] else 0
        rate_per_grams = float(rate_per_grams_list[i]) if rate_per_grams_list[i] else 0

        # Add the data row to the generated table
        table_row = {
            "rm": rm_list[i],
            "qty_pcs": qty_pcs,
            "met_wt_gms": met_wt_gms,
            "value_usd": value_usd,
            "rate_per_grams": rate_per_grams,
        }
        
        generated_table_data.append(table_row)

        # Accumulate totals
        total_qty_pcs += qty_pcs
        total_met_wt_gms += met_wt_gms
        total_value_usd += value_usd

    # Add the total row to the generated table data
    total_row_for_chalan = {
        "rm": "TOTAL",
        "qty_pcs": total_qty_pcs,
        "met_wt_gms": total_met_wt_gms,
        "value_usd": total_value_usd,
        "rate_per_grams": None,  # Set to `None` if no total for this column is required
    }
    generated_table_data.append(total_row_for_chalan)

    # Debug: Print generated table data
    # print("Generated Table Data:", generated_table_data)

    # Convert table data to JSON
    table_data_json = json.dumps(generated_table_data)

    # Store the table data in the `generated_tables` table
    cur = mysql.connection.cursor()
    insert_generated_table_query = """
        INSERT INTO generated_tables (challan_no, data)
        VALUES (%s, %s)
        ON DUPLICATE KEY UPDATE data = VALUES(data)
    """
    cur.execute(insert_generated_table_query, (challan_no_value, table_data_json))
    mysql.connection.commit()

    # Fetch existing table data for the given challan number
    cur = mysql.connection.cursor()
    select_generated_table_query = "SELECT data FROM generated_tables WHERE challan_no = %s"
    cur.execute(select_generated_table_query, (challan_no_value,))
    result = cur.fetchone()

    if result:
        # If data exists, load it and generate the table
        stored_table_data = json.loads(result[0])

        # Debug: Print stored table data
        # print("Stored Table Data:", stored_table_data)

        # Use the stored data to generate the table in Excel
        for i, row in enumerate(stored_table_data):
            ws.cell(row=data_start_row + i, column=1, value=row['rm']).alignment = LEFT_ALIGN
            if qty_pcs_exists:
                ws.cell(row=data_start_row + i, column=2, value=row['qty_pcs']).alignment = LEFT_ALIGN
            ws.cell(row=data_start_row + i, column=met_wt_gms_col, value=row['met_wt_gms']).alignment = LEFT_ALIGN
            ws.cell(row=data_start_row + i, column=value_usd_col, value=row['value_usd']).alignment = LEFT_ALIGN

            # Add logic for total row formatting
            if row['rm'] == "TOTAL":
                ws.cell(row=data_start_row + i, column=1).font = Font(bold=True)
                ws.cell(row=data_start_row + i, column=2).font = Font(bold=True)
                ws.cell(row=data_start_row + i, column=met_wt_gms_col).font = Font(bold=True)
                ws.cell(row=data_start_row + i, column=value_usd_col).font = Font(bold=True)
    else:
        print("No existing data found for challan number:", challan_no_value)

    
   # Form input
    rm_list_for_present_ppl = request.form.getlist('rm[]')  # Ensure it's a list
    rm_list_for_present_ppl = [rm.strip() for rm in rm_list_for_present_ppl]  # Strip any whitespace

    # Calculate row positions
    last_data_row_of_chalan = last_data_row + 2  # Last row where chalan ends
    headers_row_for_present_ppl = last_data_row_of_chalan + 6  # Header row
    data_start_row_for_present_ppl = headers_row_for_present_ppl + 1  # Row after headers

     # Get the input value for Challan No.
    challan_no_value = request.form.get('challan_no', '')
    
    required_columns_for_reco = ['Inv Rate','Inv Pure Wt','Inv Rm Wt','Inv Value']
    # Check for missing columns and add them with default value of 0
    for col in required_columns_for_reco:
        if col not in df.columns:
            df[col] = 0  # Add the missing column with default value 0
    
    
    # Ensure lengths match
    if len(rate_per_grams_list) < len(df):
        # Extend the list with default values (e.g., 0) if it's shorter
        rate_per_grams_list.extend([0] * (len(df) - len(rate_per_grams_list)))
    elif len(rate_per_grams_list) > len(df):
        # Trim the list if it's longer
        rate_per_grams_list = rate_per_grams_list[:len(df)]

    df['Rate'] = [float(rate) if rate else 0 for rate in rate_per_grams_list]

    # Ensure 'Inv Rate' is numeric and rounded
    df['Inv Rate'] = pd.to_numeric(df['Inv Rate'], errors='coerce').round(3)

    # Convert rate_per_grams_list to a list of rounded float values
    rate_list = [round(float(rate), 3) for rate in rate_per_grams_list if rate]


        # Check if rate_list is valid and not empty
    if rate_per_grams_list:
        # Filter rows where 'Inv Rate' matches any value in rate_list
        filtered_df = df[df['Inv Rate'].isin(rate_list)]

        # Check the filtered DataFrame
        # print("Filtered DataFrame:")
        # print(filtered_df)
        # If there are matching rows, perform groupby
        if not filtered_df.empty:
            group_for_reconciliation = (
                filtered_df.groupby(["Inv Rate"], dropna=False)
                .agg({
                    "Inv Rm Wt": "sum",
                    "Inv Pure Wt": "sum",
                    "Inv Value": "sum"
                })
                .reset_index()
            )

            group_for_reconciliation[["Inv Rate", "Inv Rm Wt", "Inv Pure Wt", "Inv Value"]] = (
                group_for_reconciliation[["Inv Rate", "Inv Rm Wt", "Inv Pure Wt", "Inv Value"]].round(3)
            )
        else:
            # Create an empty DataFrame with the necessary columns
            group_for_reconciliation = pd.DataFrame(columns=["Inv Rate", "Inv Rm Wt", "Inv Pure Wt", "Inv Value"])
    
    # Reindex with all rates from rate_per_grams_list to ensure proper order
    ordered_rates = pd.DataFrame({'Inv Rate': rate_list})

    # Merge on 'Inv Rate' ensuring both columns are of the same type and rounding
    group_for_reconciliation['Inv Rate'] = pd.to_numeric(group_for_reconciliation['Inv Rate'], errors='coerce').round(3)
    ordered_rates['Inv Rate'] = pd.to_numeric(ordered_rates['Inv Rate'], errors='coerce').round(3)

    # Perform the merge (with 'left' join to retain all rows from ordered_rates)
    group_for_reconciliation = ordered_rates.merge(
        group_for_reconciliation,
        on="Inv Rate",
        how="left"
    )
    # Fill missing values with empty spaces or zeros as needed
    group_for_reconciliation[["Inv Rm Wt", "Inv Pure Wt", "Inv Value"]] = (
        group_for_reconciliation[["Inv Rm Wt", "Inv Pure Wt", "Inv Value"]].fillna("")
    )
    
    # Debugging output
    # print("Final Group for Reconciliation:")
    # print(group_for_reconciliation)
        
    # # Add "Total" row
    # total_row_index = len(rm_list_for_present_ppl)  # Position for "Total" row 
            
    # Get the input without forcing it into a string
    banker_detail_value = request.form.get('Banker_details', '')

    # Print the banker header
    ws['E26'] = "Banker :"
    ws['E26'].font = Font(bold=True)
    ws['E26'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)

    # Write the banker details below the header
    ws['E27'] = banker_detail_value
    ws['E27'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)

    # ** Concatenate RM values with commas and write in a single cell ** 
    rm_values_string = ', '.join([str(rm) for rm in rm_list])  # Join all RM values with commas
    ws['E27'] = f"{banker_detail_value} {rm_values_string}"  # Combine Banker details with RM values
    ws['E27'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)

    # Add "Challan No." header at row 29, column E
    ws['E29'] = "Challan No."
    ws['E29'].font = Font(bold=True)  # Make the header bold
    ws['E29'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)  # Align left and prevent wrap

    # Print the value of the Challan No. in the next column (F29)
    ws['F29'] = challan_no_value  # Assign the Challan No. value
    ws['F29'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)  # Align left and prevent wrap
    
    # Get the input value for request_id .
    request_id = request.form.get('request_id', '')

    # Add "request_id." header at row 28, column E
    ws['E28'] = "Request id."
    ws['E28'].font = Font(bold=True)  # Make the header bold
    ws['E28'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)  # Align left and prevent wrap

    # Print the value of the request_id . in the next column (F28)
    ws['F28'] = request_id  # Assign the Challan No. value
    ws['F28'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)  # Align left and prevent wrap
    

    # Calculate where to print the exchange rate
    # last_data_row_of_present_ppl = data_start_row_for_present_ppl + total_row_index + 5
   
   
    # Step 1: Identify all "Pure Wt" columns
    pure_wt_columns = [col for col in final_output.columns if 'pure wt' in col.lower()]

    # Step 2: Exclude "Pure Wt Silver" columns
    pure_wt_gold_columns = [col for col in pure_wt_columns if 'silver' not in col.lower()]

    # Step 3: Check if we found any relevant "Pure Wt" columns
    if pure_wt_gold_columns:
        # Step 4: Sum all non-silver "Pure Wt" columns row-wise for 0.995 Gold
        final_output["pure wt (gms) 0.995 gold"] = final_output[pure_wt_gold_columns].sum(axis=1)

       # Convert the column to numeric, forcing errors to NaN
        final_output["pure wt (gms) 0.995 gold"] = pd.to_numeric(final_output["pure wt (gms) 0.995 gold"], errors='coerce')

        # Calculate the sum again
        column_e_sum = final_output["pure wt (gms) 0.995 gold"].sum()
        # Step 5: Place the summed values in the correct Excel row (Column E)
        for row_idx, value in enumerate(final_output["pure wt (gms) 0.995 gold"], start=start_row):
            set_cell(ws, f'E{row_idx}', value, alignment=CENTER_ALIGN)

        # print(f"Pure weight summation for 0.995 Gold added successfully in column E. Total sum: {column_e_sum}")
        # print(final_output.columns)
        # print(final_output["pure wt (gms) 0.995 gold"].head())

    else:
        print("No 'Pure Wt' columns found for 0.995 Gold.")

       
    # Convert "gross wt (gms)" to numeric, forcing errors to NaN (if any)
    final_output["gross wt (gms)"] = pd.to_numeric(final_output["gross wt (gms)"], errors='coerce')

    # Now recalculate numeric_columns after conversion
    numeric_columns = final_output.select_dtypes(include=['number']).columns

    # Calculate the sum for each numeric column
    totals_row = {col: final_output[col].sum() for col in numeric_columns}

    # Add the specific sum for "Pure Wt (gms) 0.995 Gold" into column E
    totals_row["pure wt (gms) 0.995 gold"] = column_e_sum
  
    # Add a label for the totals row (e.g., in the first column)
    totals_row.update({col: None for col in final_output.columns if col not in numeric_columns})
    totals_row["gross wt (gms)"] = round(final_output["gross wt (gms)"].sum(), 3)

    # Append the totals row to the DataFrame
    final_output_total = pd.concat([final_output, pd.DataFrame([totals_row])], ignore_index=True)    
    last_row_of_total = final_output_total.iloc[-1]

    try:
        # ========= Step A: Process the Excel File for Filtered Qty Values =========
        df_original = pd.read_excel(input_file, header=3)
        df_original.columns = df_original.columns.str.strip()  # Remove extra whitespace
        print("Columns found:", df_original.columns.tolist())

        # Remove rows where the 'Qty' column is blank (after stripping whitespace)
        df_filter = df_original[df_original["Qty"].notna() & (df_original["Qty"].astype(str).str.strip() != "")]

        # Group by 'Karatage' and sum the 'Qty' values (order preserved)
        grouped_qty = df_filter.groupby("Karatage", sort=False, as_index=False).agg({"Qty": "sum"})
        filtered_qty_values = grouped_qty["Qty"].tolist()
        print("Filtered Qty Values (grouped by Karatage & summed):", filtered_qty_values)


        # ========= Step B: Get Form Data and Setup =========
        challan_no_value = request.form.get('challan_no', '').strip()
        if not challan_no_value:
            return jsonify({'error': 'Challan number is required'})

        rm_list_for_present_ppl = request.form.getlist('rm[]')
        if not rm_list_for_present_ppl:
            return jsonify({'error': 'RM list is required'})

        invoice_no_date = request.form.get('invoice_no_date', '').strip()
        if not invoice_no_date:
            return jsonify({'error': 'Invoice number date is required'})

        if len(rm_list_for_present_ppl) != len(group_for_reconciliation):
            return jsonify({'error': 'RM list length does not match reconciliation data rows'})

        cur = mysql.connection.cursor()

        # ========= Step C: Get/Update Challan and Create a New Batch =========
        select_challan_query = "SELECT challan_id FROM challan WHERE challan_no = %s"
        cur.execute(select_challan_query, (challan_no_value,))
        challan = cur.fetchone()

        if not challan:
            insert_challan_query = "INSERT INTO challan (challan_no, invoice_no_date) VALUES (%s, %s)"
            cur.execute(insert_challan_query, (challan_no_value, invoice_no_date))
            mysql.connection.commit()
            challan_id = cur.lastrowid
        else:
            challan_id = challan[0]
            update_challan_query = "UPDATE challan SET invoice_no_date = %s WHERE challan_id = %s"
            cur.execute(update_challan_query, (invoice_no_date, challan_id))
            mysql.connection.commit()

        insert_batch_query = "INSERT INTO batch (challan_id, invoice_no_date) VALUES (%s, %s)"
        cur.execute(insert_batch_query, (challan_id, invoice_no_date))
        mysql.connection.commit()
        batch_id = cur.lastrowid

        # ========= Step D: Prepare the Mapping for Qty PCS =========
        # Here we determine which positions in the original form (qty_pcs_list) had a value.
        chalan_qty_positions = [i for i, qty in enumerate(qty_pcs_list) if qty.strip() != ""]
        mapping_index_to_qty = {}
        for pos, qty_val in zip(chalan_qty_positions, filtered_qty_values):
            mapping_index_to_qty[pos] = qty_val

        print("Mapping from chalan index to summed Qty:", mapping_index_to_qty)

        # ========= Step E: Insert Data for This Batch =========
        generated_table = group_for_reconciliation.to_dict(orient='records')

        # Validate and sanitize generated_table values
        for i, row in enumerate(generated_table):
            try:
                row["Inv Rm Wt"] = float(row["Inv Rm Wt"]) if str(row["Inv Rm Wt"]).strip() else 0.0
                row["Inv Value"] = float(row["Inv Value"]) if str(row["Inv Value"]).strip() else 0.0
                row["Inv Pure Wt"] = float(row["Inv Pure Wt"]) if str(row["Inv Pure Wt"]).strip() else 0.0
            except ValueError as e:
                return jsonify({'error': f"Invalid data in row {i+1}: {e}"})

        data_to_insert = []
        for i, row in enumerate(generated_table):
            rm_name = rm_list_for_present_ppl[i]
            if rm_name == "0.995 Gold":
                met_wt_gms = row["Inv Pure Wt"]
            else:
                met_wt_gms = row["Inv Rm Wt"]
            # Use the filtered value (if available) rather than the raw form input
            qty_pcs_value = mapping_index_to_qty.get(i, 0)
            data_to_insert.append((
                challan_id,
                batch_id,
                rm_name,
                met_wt_gms,
                row["Inv Value"],
                qty_pcs_value  # Insert the filtered qty_pcs value
            ))

        insert_reconciliation_query = """
            INSERT INTO reconciliation (challan_id, batch_id, rm, met_wt_gms, value_usd, qty_pcs)
            VALUES (%s, %s, %s, %s, %s, %s)
        """
        cur.executemany(insert_reconciliation_query, data_to_insert)
        mysql.connection.commit()

        # ========= Step F: Batch Table Excel Output =========
        # (This part uses the inserted reconciliation data to create the batch table output in Excel)
        fetch_batches_query = """
            SELECT batch_id, DATE_FORMAT(created_at, '%%Y-%%m-%%d %%H:%%i:%%s') AS batch_time, invoice_no_date
            FROM batch WHERE challan_id = %s
            ORDER BY created_at ASC
        """
        cur.execute(fetch_batches_query, (challan_id,))
        batches = cur.fetchall()

        current_row = data_start_row_for_present_ppl + 5
        qty_pcs_present = any("qty_pcs" in row for row in generated_table_data)
        if qty_pcs_present:
            headers = ["RM", "Qty Pcs", "Met Wt Gms", "Value USD"]
            merge_end_col = 4
        else:
            headers = ["RM", "Met Wt Gms", "Value USD"]
            merge_end_col = 3

        global_rec_index = 0
        for batch in batches:
            batch_id, batch_time, invoice_no_date = batch
            fetch_records_query = """
                SELECT rm, met_wt_gms, value_usd, qty_pcs
                FROM reconciliation WHERE batch_id = %s
            """
            cur.execute(fetch_records_query, (batch_id,))
            rows = cur.fetchall()

            ws.cell(row=current_row, column=1, value=f"Less: Metal Used In Packing List {invoice_no_date}")
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=merge_end_col)
            current_row += 1

            for col_num, header in enumerate(headers, 1):
                ws.cell(row=current_row, column=col_num, value=header)
            current_row += 1

            total_met_wt_gms = 0
            total_value_usd = 0
            total_qty_pcs = 0 if qty_pcs_present else None

            for db_row in rows:
                rm_name = db_row[0]
                met_wt_gms = float(db_row[1])
                value_usd = float(db_row[2])
                if qty_pcs_present:
                    qty_pcs = float(db_row[3])
                    row_values = [rm_name, qty_pcs, met_wt_gms, value_usd]
                    total_qty_pcs += qty_pcs
                else:
                    row_values = [rm_name, met_wt_gms, value_usd]

                for col_num, cell_value in enumerate(row_values, 1):
                    ws.cell(row=current_row, column=col_num, value=cell_value)
                current_row += 1

                total_met_wt_gms += met_wt_gms
                total_value_usd += value_usd
                global_rec_index += 1

            if qty_pcs_present:
                total_row_values = ["Total", total_qty_pcs, total_met_wt_gms, total_value_usd]
            else:
                total_row_values = ["Total", total_met_wt_gms, total_value_usd]

            for col_num, cell_value in enumerate(total_row_values, 1):
                ws.cell(row=current_row, column=col_num, value=cell_value)
            current_row += 2

        # ========= Step G: Compute and Output the Balance Table =========
        select_challan_data_query = """
            SELECT JSON_EXTRACT(data, '$') AS challan_data 
            FROM generated_tables 
            WHERE challan_no = %s
        """
        cur.execute(select_challan_data_query, (challan_no_value,))
        challan_data_result = cur.fetchone()
        if not challan_data_result:
            return jsonify({'error': 'No challan data found for the provided challan number.'})
        challan_data = json.loads(challan_data_result[0])
        qty_pcs_present = any(row.get('qty_pcs') is not None for row in challan_data)

        challan_map = {}
        for row in challan_data:
            rm = row['rm']
            challan_map[rm] = {
                'met_wt_gms': row.get('met_wt_gms', 0),
                'value_usd': row.get('value_usd', 0)
            }
            if qty_pcs_present:
                challan_map[rm]['qty_pcs'] = row.get('qty_pcs', 0)

        aggregate_batch_query = """
            SELECT rm, 
                SUM(met_wt_gms) AS total_met_wt_gms, 
                SUM(value_usd) AS total_value_usd,
                SUM(qty_pcs) AS total_qty_pcs
            FROM reconciliation 
            WHERE challan_id = %s 
            GROUP BY rm
        """
        cur.execute(aggregate_batch_query, (challan_id,))
        batch_data = cur.fetchall()
        batch_map = {
            row[0]: {
                'total_met_wt_gms': row[1] or 0,
                'total_value_usd': row[2] or 0,
                'total_qty_pcs': row[3] or 0
            }
            for row in batch_data
        }

        balance_table = []
        total_met_wt_gms = 0
        total_value_usd = 0
        total_qty_pcs = 0
        filtered_challan_map = {rm: values for rm, values in challan_map.items() if rm != "TOTAL"}

        for rm, challan_values in filtered_challan_map.items():
            batch_values = batch_map.get(rm, {'total_met_wt_gms': 0, 'total_value_usd': 0, 'total_qty_pcs': 0})
            balance_row = {
                'rm': rm,
                'balance_met_wt_gms': float(challan_values['met_wt_gms']) - float(batch_values['total_met_wt_gms']),
                'balance_value_usd': float(challan_values['value_usd']) - float(batch_values['total_value_usd']),
            }
            if qty_pcs_present:
                balance_row['balance_qty_pcs'] = float(challan_values.get('qty_pcs', 0)) - float(batch_values.get('total_qty_pcs', 0))
                total_qty_pcs += balance_row['balance_qty_pcs']
            balance_table.append(balance_row)
            total_met_wt_gms += balance_row['balance_met_wt_gms']
            total_value_usd += balance_row['balance_value_usd']

        total_row_for_reco = {
            'rm': 'TOTAL',
            'balance_met_wt_gms': total_met_wt_gms,
            'balance_value_usd': total_value_usd,
        }
        if qty_pcs_present:
            total_row_for_reco['balance_qty_pcs'] = total_qty_pcs
        balance_table.append(total_row_for_reco)

        current_row += 2
        if return_switch == "on":
            ws.cell(row=current_row, column=1, value="Balance Loose Metal Return")
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        else:
            ws.cell(row=current_row, column=1, value="Balance Loose Metal ")
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        current_row += 2

        if qty_pcs_present:
            balance_headers = ["RM", "Qty Pcs", "Met Wt Gms", "Value USD"]
        else:
            balance_headers = ["RM", "Met Wt Gms", "Value USD"]

        for col_num, header in enumerate(balance_headers, 1):
            ws.cell(row=current_row, column=col_num, value=header)
        current_row += 1

        for row in balance_table:
            if qty_pcs_present:
                ws.cell(row=current_row, column=1, value=row['rm'])
                ws.cell(row=current_row, column=2, value=abs(row.get('balance_qty_pcs', 0)))
                ws.cell(row=current_row, column=3, value=abs(row['balance_met_wt_gms']))
                ws.cell(row=current_row, column=4, value=abs(row['balance_value_usd']))
            else:
                ws.cell(row=current_row, column=1, value=row['rm'])
                ws.cell(row=current_row, column=2, value=abs(row['balance_met_wt_gms']))
                ws.cell(row=current_row, column=3, value=abs(row['balance_value_usd']))
            current_row += 1

        cur.close()


        if return_switch == "on":
            # Initialize starting row for writing data
            i = header_row_for_return + 1  # Starting row for writing data

            # Identify the last column (for value USD) via the column letter
            second_last_col_index = ws.max_column
            second_last_col_letter = get_column_letter(second_last_col_index)

            # --- Setup: Fuzzy Matching & Header Mapping ---
            def is_similar(str1, str2, threshold=0.6):
                """Return True if the similarity ratio between str1 and str2 is at least threshold."""
                return difflib.SequenceMatcher(None, str1.lower(), str2.lower()).ratio() >= threshold

            # IMPORTANT: Set header_row to the row where your desired headers exist.
            header_row = 33  # Adjust if your header row is different
            header_mapping = {}
            for col in range(1, ws.max_column + 1):
                header_val = ws.cell(row=header_row, column=col).value
                if header_val:
                    header_mapping[col] = str(header_val).strip()
            print("DEBUG: header_mapping =", header_mapping)
            print("DEBUG: rm_list =", rm_list)
            # --- End Setup ---

            # Loop through all balance rows except the last (total) row
            for balance_row in balance_table[:-1]:
                # Ensure required keys exist in the row
                if all(key in balance_row for key in ['balance_met_wt_gms', 'balance_value_usd', 'rm']):
                    print(f"Writing balance_met_wt_gms: {balance_row['balance_met_wt_gms']} and balance_value_usd: {balance_row['balance_value_usd']} for RM: {balance_row['rm']}")

                    # Write common fixed columns:
                    ws[f'D{i}'] = abs(balance_row['balance_met_wt_gms'])  # Column D for gross wt
                    ws[f'{second_last_col_letter}{i}'] = abs(balance_row['balance_value_usd'])  # Last column for value USD

                    # --- Determine Best Matching RM from rm_list using is_similar ---
                    row_rm = str(balance_row['rm']).strip().lower()
                    best_similarity = 0
                    best_user_rm = None

                    for user_rm in rm_list:
                        user_rm_norm = user_rm.strip().lower()
                        # Only consider this user_rm if it is similar enough
                        if not is_similar(user_rm_norm, row_rm, threshold=0.6):
                            continue

                        sim = difflib.SequenceMatcher(None, user_rm_norm, row_rm).ratio()
                        if sim > best_similarity:
                            best_similarity = sim
                            best_user_rm = user_rm

                    if best_similarity < 0.6 or best_user_rm is None:
                        print(f"No sufficient RM match found for row with RM '{row_rm}'")
                    else:
                        user_rm_norm = best_user_rm.strip().lower()
                        normalized_input = user_rm_norm.replace(" ", "")
                        # Determine forced header based on normalized input.
                        if "0.995" in normalized_input:
                            forced_header = "pure wt (gms) 0.995 gold"
                        elif "14ktgold" in normalized_input:
                            forced_header = "net wt (gms) 14kt gold"
                        elif "18ktgold" in normalized_input:
                            forced_header = "net wt (gms) 18kt gold"
                        elif "0.999pt" in normalized_input:
                            forced_header = "pt 0.999 pure wt"
                        else:
                            forced_header = user_rm_norm  # fallback to user input itself

                        target_header = None
                        # Search header_mapping for a header containing the forced header substring
                        for col_idx, header_text in header_mapping.items():
                            if forced_header in header_text.lower():
                                target_header = col_idx
                                break

                        if target_header is not None:
                            col_letter = get_column_letter(target_header)
                            print(f"Mapping for RM '{row_rm}' using forced header '{forced_header}': writing gross wt {abs(balance_row['balance_met_wt_gms'])} at {col_letter}{i}")
                            ws[f'{col_letter}{i}'] = abs(balance_row['balance_met_wt_gms'])
                        else:
                            # Fallback: perform fuzzy matching over header_mapping if forced mapping fails
                            best_similarity_header = 0
                            best_match_col = None
                            for col_idx, header_text in header_mapping.items():
                                sim_header = difflib.SequenceMatcher(None, user_rm_norm, header_text.lower()).ratio()
                                if sim_header > best_similarity_header:
                                    best_similarity_header = sim_header
                                    best_match_col = col_idx
                            if best_similarity_header >= 0.6 and best_match_col is not None:
                                col_letter = get_column_letter(best_match_col)
                                print(f"Fallback best match for RM '{row_rm}' and user_rm '{user_rm_norm}' is header '{header_mapping[best_match_col]}' with similarity {best_similarity_header:.2f}; writing gross wt {abs(balance_row['balance_met_wt_gms'])} at {col_letter}{i}")
                                ws[f'{col_letter}{i}'] = abs(balance_row['balance_met_wt_gms'])
                            else:
                                print(f"No header match found for RM '{row_rm}' with user input '{user_rm_norm}'")
                    # --- End Mapping ---

                    i += 1  # Move to the next row
                else:
                    print(f"Skipping row due to missing keys: {balance_row}")
        else:
            # Handle the case when return_switch is not "on" if necessary
            pass  # Or add appropriate logic

        table = current_row + 5
        exchange_rate_row_number = table + 15 # Add a 5-line gap (3 for data, 2 for space)

        # --- Step 1: Clean Column Names ---
        df.columns = df.columns.astype(str).str.strip()
        df.columns = df.columns.str.replace(r'\s+', ' ', regex=True)
        df.columns = df.columns.str.replace(r'[^A-Za-z0-9 ]+', '', regex=True)

        # --- Step 2: Ensure Required Columns Exist ---
        required_columns = ["Ctg", "Rm Code"]
        for col in required_columns:
            if col not in df.columns:
                df[col] = ""

        # --- Step 3: Force Each Value to a String Explicitly ---
        # Using .apply(lambda x: ...) is sometimes more robust than .astype(str)
        df["Ctg"] = df["Ctg"].apply(lambda x: str(x) if pd.notnull(x) else "")
        df["Rm Code"] = df["Rm Code"].apply(lambda x: str(x) if pd.notnull(x) else "")

        # --- Debug: Check the Data Types of Each Element ---
        print("Unique types in 'Ctg':", df["Ctg"].apply(type).unique())
        print("Unique types in 'Rm Code':", df["Rm Code"].apply(type).unique())

        # --- Step 4: Filter Rows Where "Ctg" Equals "C" ---
        # Now that all values should be strings, .str.upper() should work fine.
        filtered_df = df[df["Ctg"].str.upper() == "C"].copy()

        # --- Step 5: Map "Rm Code" Values Using the Dictionary ---
        # Ensure the dictionary keys are compared as strings.
        mapped_values = filtered_df["Rm Code"].apply(lambda x: constants.COLOUR_STONE_CODE.get(x, None))
        
        # Remove None values and get unique values
        mapped_values = mapped_values.dropna().unique()

        # --- Step 6: Join All Mapped Values With a Slash ---
        joined_string = "/".join(mapped_values)

        
        diamond_stone_table = df.loc[df['Ctg'].isin(['C','D'])].groupby(["Ctg"], dropna=False).agg({
            "Inv Rm Wt": "sum",
            "Inv Value": "sum",
            "Inv Rm Qty": "sum"
        }).reset_index()
    
        diamond_wt = diamond_value = diamond_qty = CS_wt = CS_value = CS_qty = 0
        for i, row in diamond_stone_table.iterrows():
            if row['Ctg'] == 'D':
                diamond_wt = round(row['Inv Rm Wt'], 3)
                diamond_value = round(row['Inv Value'],2)
                diamond_qty = (row['Inv Rm Qty'])
            elif row['Ctg'] == 'C':
                CS_wt = round(row['Inv Rm Wt'],3)
                CS_value = round(row['Inv Value'],2)
                CS_qty = (row['Inv Rm Qty'])


    # Adding "Type", "PCS", "Total CTW", "Net Payable by You"
        set_cell(ws, f'B{table}', "PCS", font=BOLD_FONT, alignment=LEFT_ALIGN)
        set_cell(ws, f'C{table}', "Total CTW", font=BOLD_FONT, alignment=LEFT_ALIGN)
        set_cell(ws, f'D{table}', "Net Payable by You", font=BOLD_FONT, alignment=LEFT_ALIGN)
        set_cell(ws, f'A{table}', "Type", font=BOLD_FONT, alignment=LEFT_ALIGN)

        # Adding data row (e.g., "Diamond")
        set_cell(ws, f'A{table + 1}', "Diamond", font=BOLD_FONT, alignment=LEFT_ALIGN)
        set_cell(ws, f'A{table + 2}', joined_string, font=BOLD_FONT, alignment=LEFT_ALIGN)  # Dynamically set mapped value

        set_cell(ws, f'A{table + 5}', "Type", font=BOLD_FONT, alignment=LEFT_ALIGN)
        set_cell(ws, f'B{table + 5}', "Value", font=BOLD_FONT, alignment=LEFT_ALIGN)

        set_cell(ws, f'A{table + 6}', "Diamond", font=BOLD_FONT, alignment=LEFT_ALIGN)
        set_cell(ws, f'A{table + 7}', joined_string, font=BOLD_FONT, alignment=LEFT_ALIGN)  # Dynamically set mapped value
        set_cell(ws, f'A{table + 8}', "Labour", font=BOLD_FONT, alignment=LEFT_ALIGN)
        set_cell(ws, f'A{table + 9}', "Grand Total", font=BOLD_FONT, alignment=LEFT_ALIGN)
        set_cell(ws, f'A{table + 3}', "Grand Total", font=BOLD_FONT, alignment=LEFT_ALIGN)
        
        target_row = table + 11  # Replace 'table' with your variable's actual value
        target_row_calculation = target_row + 1
        target_row_calculation_for_net_realization =  target_row_calculation + 1
        last_row_of_total_values = last_row_of_total.tolist()

        for col_idx, value in enumerate(last_row_of_total_values, start=1):  # start=1 for column A
            ws.cell(row=target_row, column=col_idx, value=value)
            
      
        if isinstance(last_row_of_total, pd.Series):
            # Calculate the last column index
            last_column_index = len(last_row_of_total_values) - 1  # Subtract 1 to shift the columns to the left
            last_column_letter = get_column_letter(last_column_index)  # Get the correct Excel column letter
            last_column_letter_for_metal = get_column_letter(last_column_index)  # Same adjustment for the metal column

            # Set the values for "Less: Metal Cost US$" and "Net Realization US$" in the shifted columns
            set_cell(ws, f'{last_column_letter}{target_row_calculation}', "Less: Metal Cost US$", font=BOLD_FONT, alignment=LEFT_ALIGN)
            set_cell(ws, f'{last_column_letter}{target_row_calculation_for_net_realization}', "Net Realization US$", font=BOLD_FONT, alignment=LEFT_ALIGN)

        # Continue processing metal_amt_sum and net_realization_value
        final_output.columns = final_output.columns.str.strip().str.lower()  # Make column names lowercase and strip spaces

        # Step 2: Check if 'metal amt.' column exists (case insensitive)
        metal_amt_column_name = [col for col in final_output.columns if 'metal amt' in col.lower()]
        amount_col = [col for col in final_output.columns if 'amount us$' in col.lower()]

        if metal_amt_column_name:
            # Step 3: Sum the 'Metal Amt.' column
            metal_amt_sum = final_output[metal_amt_column_name[0]].sum()  # Use the first matching column

            # Shift one column to the right
            new_column_index = column_index_from_string(last_column_letter_for_metal) + 1
            new_column_letter = get_column_letter(new_column_index)

            set_cell(ws, f'{new_column_letter}{target_row_calculation}', metal_amt_sum, font=BOLD_FONT, alignment=RIGHT_ALIGN)

        if amount_col:
            # Step 3: Sum the 'Amount US$' column
            amt_col_sum = final_output[amount_col[0]].sum()  # Use the first matching column
            net_realization_value = amt_col_sum - metal_amt_sum

            # Shift one column to the right
            new_column_index = column_index_from_string(last_column_letter_for_metal) + 1
            new_column_letter = get_column_letter(new_column_index)

            set_cell(ws, f'{new_column_letter}{target_row_calculation_for_net_realization}', net_realization_value, font=BOLD_FONT, alignment=RIGHT_ALIGN)

        if "Inv Rm Qty" in df.columns:
            ws[f'B{table + 1}'] = diamond_qty 
            ws[f'B{table + 1}'].font = BOLD_FONT
            ws[f'B{table + 1}'].alignment = LEFT_ALIGN
            ws[f'B{table + 2}'] = CS_qty
            ws[f'B{table + 2}'].font = BOLD_FONT
            ws[f'B{table + 2}'].alignment = LEFT_ALIGN
            total_qty_table = diamond_qty + CS_qty
        
            ws[f'B{table + 3}'] = total_qty_table 
            ws[f'B{table + 3}'].font = BOLD_FONT
            ws[f'B{table + 3}'].alignment = LEFT_ALIGN

        if "Inv Rm Wt" in df.columns:
            ws[f'C{table + 1}'] = diamond_wt
            ws[f'C{table + 1}'].font = BOLD_FONT
            ws[f'C{table + 1}'].alignment = LEFT_ALIGN
            ws[f'C{table + 2}'] = CS_wt 
            ws[f'C{table + 2}'].font = BOLD_FONT
            ws[f'C{table + 2}'].alignment = LEFT_ALIGN

            total_rm_wt_table = round(diamond_wt + CS_wt,3)
            ws[f'C{table + 3}'] = total_rm_wt_table 
            ws[f'C{table + 3}'].font = BOLD_FONT
            ws[f'C{table + 3}'].alignment = LEFT_ALIGN

        if "Inv Value" in df.columns:
            try:
                # Diamond values
                ws[f'D{table + 1}'] = diamond_value
                ws[f'B{table + 6}'] = diamond_value
                ws[f'D{table + 1}'].font = BOLD_FONT
                ws[f'D{table + 1}'].alignment = LEFT_ALIGN

                # Precious/Semi-Precious Stone values
                ws[f'D{table + 2}'] = CS_value
                ws[f'B{table + 7}'] = CS_value
                ws[f'D{table + 2}'].font = BOLD_FONT
                ws[f'D{table + 2}'].alignment = LEFT_ALIGN


                total_inv_value_table = round(diamond_value + CS_value,2)
                ws[f'D{table + 3}'] = total_inv_value_table 
                ws[f'D{table + 3}'].font = BOLD_FONT
                ws[f'D{table + 3}'].alignment = LEFT_ALIGN
            except Exception as e:
                print(f"Error processing 'Inv Value': {e}")
    
        df.columns = df.columns.str.strip()  # Ensure no leading/trailing spaces in column names
        if 'Labour' in df.columns:
            try:
                df['Labour'] = pd.to_numeric(df['Labour'], errors='coerce')
                labour_sum = round(df['Labour'].sum(skipna=True),2)
                ws[f'B{table + 8}'] = labour_sum
            except Exception as e:
                print(f"Error processing 'Labour': {e}")
        else:
            labour_sum = 0
            print("No 'Labour' column found. Setting labour_sum to 0.")

        # Calculate table total
        
        table_total = round(diamond_value + CS_value + labour_sum,2)
        ws[f'B{table + 9}'] = table_total  # Write total to the cell

        # Split the number into dollars and cents
        dollars = int(table_total)
        cents = round((table_total - dollars) * 100)  # Get the cents part as an integer

        # Convert the dollar part and cents part to words
        dollar_words = num2words(dollars).title()  # Convert to words and capitalize each word
        cents_words = num2words(cents).title()  # Convert cents to words

        # Create the final text in the required format
        if cents > 0:
            words = f"Total Net Realisation Dollar {dollar_words} & Cent {cents_words} Only."
        else:
            words = f"Total Net Realisation Dollar {dollar_words} Only."

        # Write the final text to the Excel file
        # ws[f'B{table + 10}'] = words  # Write the text in the next row
        
    
        exchange_rate_value = float(request.form['exchange_rate'])  
        amount_chargeable_row_number = exchange_rate_row_number + 2  # Leave 1 row after exchange rate for amount chargeable 
        line_1 = amount_chargeable_row_number + 3  # Leave 1 row after exchange rate for amount chargeable 
        line_2 = line_1 + 3  # Leave 2 row after line 1 for line 2
        line_3 = line_2 + 2  # Leave 2 row after line 1 for line 2
        line_4 = line_3 + 2  # Leave 2 row after line 1 for line 2
        line_5 = line_4 + 1 # Leave 2 row after line 1 for line 2
        line_6 = line_5 + 2 # Leave 2 row after line 1 for line 2
        line_7 = line_6 + 1 # Leave 1 row after line 1 for line 2
        line_8 = line_7 + 1 # Leave 1 row after line 1 for line 2

        # Write the Exchange Rate in the first column
        ws.cell(row=exchange_rate_row_number, column=1, value=f"Exchange Rate: {exchange_rate_value}")
        from openpyxl.styles import Font
        ws.cell(row=exchange_rate_row_number, column=1).font = Font(bold=True)  # Make the text bold
        
        # Leave one blank row and write "Amount Chargeable (in Words)" below the Exchange Rate
        ws.cell(row=amount_chargeable_row_number, column=1, value="Amount Chargeable (in Words)")
        ws.cell(row=amount_chargeable_row_number, column=3, value= words)

        ws.cell(row=amount_chargeable_row_number, column=1).font = Font(bold=True)  # Make the text bold
        ws.cell(row=amount_chargeable_row_number, column=3).font = Font(bold=True)  # Make the text bold
    
        ws.cell(row=line_1, column=1, value="""I/we hereby certify that my/our registration certificate under Goods and Services Tax Act,2017 is in force on the date on which the supply of goods / services specified in this Tax Invoice/ Consignment sale is made by me/us and that the transaction of sale covered by this tax invoice/Consignment sales has been effected by me/us & it shall be accounted for in the turnover of sales while filling of return & the due tax, if any, payable on the sale has been paid or shall paid.""")																						
        ws.cell(row=line_1, column=1).font = Font(bold=True)  # Make the text bold
    
        ws.cell(row=line_2, column=1, value="""SUPPLY TO SEZ UNIT IS UNDER ZERO RATED SUPPLY AS PER THE IGST RULE.""")
        ws.cell(row=line_2, column=1).font = Font(bold=True)  # Make the text bold
        
        ws.cell(row=line_3, column=1, value="""SUPPLY MEANT FOR EXPORT/SUPPLY TO SEZ UNIT UNDER BOND OR LETTER OF UNDERTAKING WITHOUT PAYMENT OF INTERGRATED TAX AS PER THE IGST RULE""")
        ws.cell(row=line_3, column=1).font = Font(bold=True)  # Make the text bold
        
        # Print the same data inside the provided sentence
        formatted_sentence = f"Gold & Silver received from SEZ UNIT vide invoice No.{invoice_no_date} being returned after job work"
        ws.cell(row=line_4, column=1, value=formatted_sentence) 
        ws.cell(row=line_4, column=1).font = Font(bold=True)  # Make the text bold
        
        ws.cell(row=line_5, column=1, value="""This Invoice is for only Labour & Diamond Charges""")
        ws.cell(row=line_5, column=1).font = Font(bold=True)  # Make the text bold
    
        ws.cell(row=line_6, column=1, value="Declaration")
        ws.cell(row=line_6, column=1).font = Font(bold=True)  # Make the text bold
        
        ws.cell(row=line_7, column=1, value="We declare that this Invoice shows the actual price of the")
        ws.cell(row=line_7, column=1).font = Font(bold=True)  # Make the text bold
    
        ws.cell(row=line_8, column=1, value="goods described and that all particulars are true and correct.")
        ws.cell(row=line_8, column=1).font = Font(bold=True)  # Make the text bold



        # Save the Excel file
        output_file = os.path.join(OUTPUT_DIR, "Formatted_Invoice.xlsx")
        wb.save(output_file)

        # Send the file for download
        return send_file(output_file, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name="Formatted_Invoice.xlsx")

    except Exception as e:
        import traceback
        error_message = str(e)
        error_trace = traceback.format_exc()
        print("Error Trace:", error_trace)  # Log the full traceback
        return jsonify({'error': error_message})
     
     
     
     


   
   
            
def get_metal_mapping(output_header, row, ctg, metal_kt):
    return_dict = {}
    common_column = "Inv Rm Wt"
    metals = ['G','S', 'M', 'NTL', 'STL', 'P','TTM','STL','NTL']
    metal_column = "Inv Pure Wt"
    stones = ['D', 'C']
    stone_column = "Inv Value"
    skip_headers = ['Metal Amt.'] 
    for header in output_header:
        if header in skip_headers: continue
        if constants.CATEGORY_DICT.get(ctg, '') in header.strip() or str(metal_kt) in header.strip():
            if ctg in metals:
                if ctg == 'G' and metal_kt not in header:
                    continue
                if 'pure' in header.lower():
                    return_dict[header] = round(row[metal_column], 3)
                else:
                    return_dict[header] = round(row[common_column], 3)
            elif ctg in stones:
                if 'cts' in header.lower():
                    return_dict[header] = round(row[common_column], 3)
                else:
                    return_dict[header] = round(row[stone_column], 2)

    return return_dict


def get_product_name(ctg_str, rm_code_str):
    # Step 1: Convert the comma-separated strings into lists
    ctg_list = ctg_str.split(',')  # List of Ctg values
    rm_code_list = rm_code_str.split(',')  # List of Rm Code values

    found = False
    product_name = ''
    skip_list = ['nan', 'M']

    for i, code in enumerate(ctg_list):  
        code = code.strip()  # Clean extra spaces

        if code in skip_list:
            continue

        if code == "C":  
            # If "C" is found, check the corresponding Rm Code
            if i < len(rm_code_list):  
                rm_code = rm_code_list[i].strip()
                # Fetch the mapped value from COLOUR_STONE_CODE dictionary
                mapped_value = constants.COLOUR_STONE_CODE.get(rm_code, '')
                product_name += ' ' + mapped_value  
        else:
            # Fetch from STATEMENT_DICT for non-"C" values
            product_name += ' ' + constants.STATEMENT_DICT.get(code, '')

        if code in ["D", "C"]:
            found = True

    # Remove leading/trailing spaces
    product_name = product_name.strip()

    # Step 3: If neither "D" nor "C" was found, prepend "Plain"
    if not found:
        product_name = "Plain " + product_name

    # Step 4: Append the fixed suffix and return
    return product_name + ' Jewellery'



def map_headers_to_data(headers, data):
    mapped_data = []
    for row in data:
        mapped_row = {header.strip(): row.get(header.strip(), None) for header in headers}
        mapped_data.append(mapped_row)
    return mapped_data


if __name__ == '__main__':
    app_job_work.run(debug=True) 



																						
